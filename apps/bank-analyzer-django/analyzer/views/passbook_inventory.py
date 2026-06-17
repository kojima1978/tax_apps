"""通帳有無一覧表ビュー"""
import json
import logging
from datetime import date
from io import BytesIO

import pandas as pd
from django.contrib import messages
from django.db.models import Min, Max
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

from ..models import Account, Case, Transaction
from ..lib.constants import ERAS
from ..templatetags.japanese_date import get_japanese_era, wareki as wareki_func
from ._helpers import sanitize_filename, set_download_filename

logger = logging.getLogger(__name__)

CERTIFICATE_IMPORT_COLUMNS = {
    'bank_name': ('銀行名', '銀行', '金融機関', '金融機関名'),
    'branch_name': ('支店名', '支店', '店舗名'),
    'account_type': ('種類', '種別', '口座種別'),
    'account_number': ('口座番号', '口座No', '口座No.', '口座'),
    'certificate_balance': ('残証残高', '残高証明残高', '残高証明書残高', '証明残高', '残高証明書'),
    'passbook_balance': ('通帳残高', '帳簿残高'),
    'has_accrued_interest': ('既経過利息', '既経過利息計算'),
    'inventory_remarks': ('備考', 'メモ', '摘要'),
}


# ---------------------------------------------------------------------------
# ヘルパー
# ---------------------------------------------------------------------------

def _get_year_range(case: Case) -> list[int]:
    """案件の取引データから年の範囲を取得"""
    agg = Transaction.objects.filter(case=case, date__isnull=False).aggregate(
        min_date=Min('date'), max_date=Max('date'),
    )
    if not agg['min_date']:
        return []
    return list(range(agg['min_date'].year, agg['max_date'].year + 1))


def _get_passbook_balance(account: Account, reference_date: date | None) -> int | None:
    """基準日以前で最も近い取引の残高を取得"""
    qs = account.transactions.filter(balance__isnull=False)
    if reference_date:
        qs = qs.filter(date__lte=reference_date)
    tx = qs.order_by('-date', '-id').first()
    return tx.balance if tx else None


def _get_transaction_years(account: Account) -> set[int]:
    """口座に取引がある年のセットを取得"""
    dates = account.transactions.filter(date__isnull=False).values_list('date', flat=True)
    return {d.year for d in dates}


def _wareki_abbr(year: int) -> str:
    """西暦年 → 和暦略称 (例: 2024 → 'R6')"""
    d = date(year, 1, 1)
    for era_start, _, abbr in ERAS:
        if d >= era_start:
            era_year = year - era_start.year + 1
            return f"{abbr}{era_year}"
    return str(year)


def _balance_match_status(passbook_bal, certificate_bal) -> str:
    """残高一致ステータスを判定"""
    if certificate_bal is None:
        return "残高証明なし"
    if passbook_bal is not None and passbook_bal == certificate_bal:
        return "○"
    if passbook_bal is not None and passbook_bal != certificate_bal:
        return "×"
    return "証明のみ"


def _parse_amount(value) -> int | None:
    """フォーム/取込値から金額をパース"""
    if value is None:
        return None
    if pd.isna(value):
        return None
    if isinstance(value, str):
        value = value.strip().replace(',', '')
        if not value:
            return None
    return int(float(value))


def _normalize_account_number(value) -> str:
    """Excel数値セルも考慮して口座番号を文字列化"""
    if value is None or pd.isna(value):
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_bool(value) -> bool:
    """取込値を真偽値に変換"""
    if value is None or pd.isna(value):
        return False
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    return text in ('1', 'true', 'yes', 'y', '有', 'あり', '○', '〇', '済')


def _get_import_value(row: dict, field: str, default=''):
    """別名を許容して取込行から値を取得"""
    for col in CERTIFICATE_IMPORT_COLUMNS[field]:
        if col in row and not pd.isna(row[col]):
            return row[col]
    return default


def _next_print_order(case: Case) -> int:
    """新規口座の末尾表示順を取得"""
    current = case.accounts.aggregate(max_order=Max('print_order')).get('max_order')
    return (current or 0) + 1


def _upsert_certificate_account(case: Case, data: dict) -> tuple[Account, bool]:
    """残高証明書ベースの口座を作成/更新"""
    account_number = _normalize_account_number(data.get('account_number'))
    if not account_number:
        raise ValueError('口座番号が必要です')

    defaults = {
        'bank_name': data.get('bank_name') or '',
        'branch_name': data.get('branch_name') or '',
        'account_type': data.get('account_type') or '',
        'certificate_balance': data.get('certificate_balance'),
        'passbook_balance': data.get('passbook_balance'),
        'has_accrued_interest': bool(data.get('has_accrued_interest')),
        'inventory_remarks': data.get('inventory_remarks') or '取引履歴なし・残高証明書あり',
        'print_order': _next_print_order(case),
    }
    account, created = Account.objects.get_or_create(
        case=case,
        account_number=account_number,
        defaults=defaults,
    )
    if created:
        return account, True

    update_fields = []
    for field in (
        'bank_name', 'branch_name', 'account_type', 'certificate_balance',
        'passbook_balance', 'has_accrued_interest', 'inventory_remarks',
    ):
        value = defaults[field]
        if value not in (None, '') or field == 'has_accrued_interest':
            setattr(account, field, value)
            update_fields.append(field)
    if update_fields:
        account.save(update_fields=update_fields)
    return account, False


def _read_certificate_import(file_obj) -> pd.DataFrame:
    """CSV/Excelの残高証明書口座リストを読み込む"""
    name = (file_obj.name or '').lower()
    content = file_obj.read()
    if name.endswith(('.xlsx', '.xls')):
        return pd.read_excel(BytesIO(content), dtype=str)

    for encoding in ('utf-8-sig', 'cp932', 'utf-8'):
        try:
            return pd.read_csv(BytesIO(content), encoding=encoding, dtype=str)
        except UnicodeDecodeError:
            continue
    return pd.read_csv(BytesIO(content), encoding='cp932', encoding_errors='replace', dtype=str)


def _import_certificate_accounts(case: Case, file_obj) -> tuple[int, int, int]:
    """残高証明書口座リストを取込"""
    df = _read_certificate_import(file_obj).fillna('')
    created_count = updated_count = skipped_count = 0

    for _, row_obj in df.iterrows():
        row = row_obj.to_dict()
        account_number = _normalize_account_number(_get_import_value(row, 'account_number'))
        if not account_number:
            skipped_count += 1
            continue

        data = {
            'bank_name': str(_get_import_value(row, 'bank_name')).strip(),
            'branch_name': str(_get_import_value(row, 'branch_name')).strip(),
            'account_type': str(_get_import_value(row, 'account_type')).strip(),
            'account_number': account_number,
            'certificate_balance': _parse_amount(_get_import_value(row, 'certificate_balance', None)),
            'passbook_balance': _parse_amount(_get_import_value(row, 'passbook_balance', None)),
            'has_accrued_interest': _parse_bool(_get_import_value(row, 'has_accrued_interest', False)),
            'inventory_remarks': str(_get_import_value(row, 'inventory_remarks', '取引履歴なし・残高証明書あり')).strip(),
        }
        _, created = _upsert_certificate_account(case, data)
        if created:
            created_count += 1
        else:
            updated_count += 1

    return created_count, updated_count, skipped_count


def _build_account_rows(case: Case, years: list[int]) -> list[dict]:
    """口座ごとの一覧データを構築"""
    accounts = case.accounts.all().order_by('print_order', 'bank_name', 'branch_name', 'account_number')
    rows = []
    for acc in accounts:
        tx_years = _get_transaction_years(acc)
        auto_balance = _get_passbook_balance(acc, case.reference_date)

        passbook_bal = acc.passbook_balance if acc.passbook_balance is not None else auto_balance
        year_list = []
        for y in years:
            saved = acc.passbook_years.get(str(y))
            has = bool(saved) if saved is not None else (y in tx_years)
            year_list.append({'year': y, 'has': has})

        rows.append({
            'id': acc.id,
            'bank_name': acc.bank_name or '',
            'branch_name': acc.branch_name or '',
            'account_type': acc.account_type or '',
            'account_number': acc.account_number or '',
            'year_list': year_list,
            'passbook_balance': passbook_bal,
            'auto_balance': auto_balance,
            'certificate_balance': acc.certificate_balance,
            'balance_match': _balance_match_status(passbook_bal, acc.certificate_balance),
            'has_accrued_interest': acc.has_accrued_interest,
            'inventory_remarks': acc.inventory_remarks,
        })
    return rows


# ---------------------------------------------------------------------------
# ビュー
# ---------------------------------------------------------------------------

def passbook_inventory(request: HttpRequest, pk: int) -> HttpResponse:
    """通帳有無一覧表ページ"""
    case = get_object_or_404(Case, pk=pk)
    years = _get_year_range(case)
    rows = _build_account_rows(case, years)

    year_labels = [{'year': y, 'wareki': f"({_wareki_abbr(y)})"} for y in years]
    total_passbook = sum(r['passbook_balance'] or 0 for r in rows)
    total_certificate = sum(r['certificate_balance'] or 0 for r in rows)

    context = {
        'case': case,
        'years': year_labels,
        'rows': rows,
        'total_passbook': total_passbook,
        'total_certificate': total_certificate,
    }
    return render(request, 'analyzer/passbook_inventory.html', context)


def add_certificate_account(request: HttpRequest, pk: int) -> HttpResponse:
    """取引履歴がない残高証明書口座を追加"""
    case = get_object_or_404(Case, pk=pk)
    if request.method != 'POST':
        return redirect('passbook-inventory', pk=pk)

    try:
        _, created = _upsert_certificate_account(case, {
            'bank_name': request.POST.get('bank_name', '').strip(),
            'branch_name': request.POST.get('branch_name', '').strip(),
            'account_type': request.POST.get('account_type', '').strip(),
            'account_number': request.POST.get('account_number', '').strip(),
            'certificate_balance': _parse_amount(request.POST.get('certificate_balance')),
            'passbook_balance': _parse_amount(request.POST.get('passbook_balance')),
            'has_accrued_interest': request.POST.get('has_accrued_interest') == 'on',
            'inventory_remarks': request.POST.get('inventory_remarks', '').strip()
                or '取引履歴なし・残高証明書あり',
        })
    except (TypeError, ValueError) as exc:
        messages.error(request, f'口座を追加できませんでした: {exc}')
        return redirect('passbook-inventory', pk=pk)

    verb = '追加' if created else '更新'
    messages.success(request, f'残高証明書の口座を{verb}しました。')
    return redirect('passbook-inventory', pk=pk)


def import_certificate_accounts(request: HttpRequest, pk: int) -> HttpResponse:
    """口座・残証残高リストをCSV/Excelから取込"""
    case = get_object_or_404(Case, pk=pk)
    if request.method != 'POST':
        return redirect('passbook-inventory', pk=pk)

    file_obj = request.FILES.get('certificate_file')
    if not file_obj:
        messages.error(request, '取込ファイルを選択してください。')
        return redirect('passbook-inventory', pk=pk)

    try:
        created, updated, skipped = _import_certificate_accounts(case, file_obj)
    except Exception as exc:
        logger.exception('残高証明書口座リスト取込エラー: case_id=%s', pk)
        messages.error(request, f'取込に失敗しました: {exc}')
        return redirect('passbook-inventory', pk=pk)

    messages.success(
        request,
        f'残高証明書の口座リストを取り込みました。追加{created}件、更新{updated}件、スキップ{skipped}件。',
    )
    return redirect('passbook-inventory', pk=pk)


def api_save_passbook_inventory(request: HttpRequest, pk: int) -> JsonResponse:
    """通帳有無一覧表のデータを保存（AJAX）"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    case = get_object_or_404(Case, pk=pk)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    account_id = data.get('account_id')
    if not account_id:
        return JsonResponse({'error': 'account_id required'}, status=400)

    account = get_object_or_404(Account, pk=account_id, case=case)

    field = data.get('field')
    value = data.get('value')

    if field == 'passbook_balance':
        account.passbook_balance = int(value) if value not in (None, '') else None
    elif field == 'certificate_balance':
        account.certificate_balance = int(value) if value not in (None, '') else None
    elif field == 'has_accrued_interest':
        account.has_accrued_interest = bool(value)
    elif field == 'inventory_remarks':
        account.inventory_remarks = value or ''
    elif field == 'passbook_year':
        year = str(data.get('year'))
        account.passbook_years[year] = bool(value)
    else:
        return JsonResponse({'error': f'Unknown field: {field}'}, status=400)

    account.save()

    passbook_bal = account.passbook_balance
    if passbook_bal is None:
        passbook_bal = _get_passbook_balance(account, case.reference_date)
    balance_match = _balance_match_status(passbook_bal, account.certificate_balance)

    return JsonResponse({'ok': True, 'balance_match': balance_match})


# ---------------------------------------------------------------------------
# Excel出力
# ---------------------------------------------------------------------------

def export_passbook_inventory(request: HttpRequest, pk: int) -> HttpResponse:
    """通帳有無一覧表をExcel出力"""
    case = get_object_or_404(Case, pk=pk)
    years = _get_year_range(case)
    rows = _build_account_rows(case, years)

    wb = Workbook()
    ws = wb.active
    ws.title = '通帳有無一覧表'

    # スタイル定義
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    peach_fill = PatternFill(start_color='FFDAB9', end_color='FFDAB9', fill_type='solid')
    hdr_font = Font(name='游ゴシック', size=9, bold=True)
    body_font = Font(name='游ゴシック', size=9)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_a = Alignment(horizontal='right', vertical='center')

    year_count = len(years)
    # Col A=No, B=銀行名, C=支店名, D=種類, E=口座番号, F..=年, 残高系5列
    COL_BANK = 2
    COL_YEAR_START = 6
    COL_AFTER = COL_YEAR_START + year_count  # 通帳残高の列
    total_cols = COL_AFTER + 4  # +残高一致+残証残高+既経過利息+備考

    # --- 行1: タイトル ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    c = ws.cell(row=1, column=1, value=f'{case.name}  通帳有無一覧表')
    c.font = Font(name='游ゴシック', size=14, bold=True)

    ref_str = wareki_func(case.reference_date, 'full') if case.reference_date else '○年○月○日'
    mid_col = COL_YEAR_START + year_count // 2
    ws.merge_cells(start_row=1, start_column=mid_col, end_row=1, end_column=COL_AFTER + 2)
    c = ws.cell(row=1, column=mid_col, value=f'相続開始日：{ref_str}')
    c.font = Font(name='游ゴシック', size=12)
    c.alignment = Alignment(horizontal='center', vertical='center')

    c = ws.cell(row=1, column=total_cols, value=f'{date.today().strftime("%Y/%m/%d")}\n(作成日)')
    c.font = Font(name='游ゴシック', size=8)
    c.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

    # --- 行2-3: ヘッダー ---
    fixed_headers = [('', 'No'), ('銀行名', None), ('支店名', None), ('種類', None), ('口座番号', None)]
    # 固定列（2行結合）
    for i, (label, _) in enumerate(fixed_headers):
        col = i + 1
        val = label or _
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
        c = ws.cell(row=2, column=col, value=val)
        c.font = hdr_font
        c.alignment = center
        c.border = border
        ws.cell(row=3, column=col).border = border

    # 年列（行2=西暦, 行3=和暦）
    for i, y in enumerate(years):
        col = COL_YEAR_START + i
        c = ws.cell(row=2, column=col, value=str(y))
        c.font = hdr_font
        c.alignment = center
        c.border = border
        c = ws.cell(row=3, column=col, value=f'({_wareki_abbr(y)})')
        c.font = hdr_font
        c.alignment = center
        c.border = border

    # 後半固定列（2行結合）
    after_headers = ['通帳\n残高', '残高\n一致', '残証\n残高', '既経過\n利息', '備考']
    for i, label in enumerate(after_headers):
        col = COL_AFTER + i
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
        c = ws.cell(row=2, column=col, value=label)
        c.font = hdr_font
        c.alignment = center
        c.border = border
        ws.cell(row=3, column=col).border = border

    # --- データ行 ---
    DATA_START = 4
    MAX_ROWS = max(12, len(rows))

    for row_i in range(MAX_ROWS):
        r = DATA_START + row_i
        has_data = row_i < len(rows)
        d = rows[row_i] if has_data else None

        # No
        c = ws.cell(row=r, column=1, value=row_i + 1)
        c.font = body_font
        c.alignment = center
        c.border = border

        # 銀行名〜口座番号
        info = [d['bank_name'], d['branch_name'], d['account_type'], d['account_number']] if d else ['', '', '', '']
        for i, val in enumerate(info):
            c = ws.cell(row=r, column=COL_BANK + i, value=val)
            c.font = body_font
            c.alignment = left_wrap
            c.border = border
            c.fill = peach_fill

        # 年列
        for i, y in enumerate(years):
            col = COL_YEAR_START + i
            val = ''
            if d and d['year_list'][i]['has']:
                val = '○'
            c = ws.cell(row=r, column=col, value=val)
            c.font = body_font
            c.alignment = center
            c.border = border
            c.fill = peach_fill

        # 通帳残高
        c = ws.cell(row=r, column=COL_AFTER, value=d['passbook_balance'] if d else None)
        c.font = body_font
        c.alignment = right_a
        c.border = border
        c.number_format = '#,##0'

        # 残高一致
        match_val = d['balance_match'] if d else '残高証明なし'
        c = ws.cell(row=r, column=COL_AFTER + 1, value=match_val)
        c.font = body_font
        c.alignment = center
        c.border = border

        # 残証残高
        c = ws.cell(row=r, column=COL_AFTER + 2, value=d['certificate_balance'] if d else None)
        c.font = body_font
        c.alignment = right_a
        c.border = border
        c.number_format = '#,##0'

        # 既経過利息計算
        if d:
            interest = '☑ 有' if d['has_accrued_interest'] else '□ 有'
        else:
            interest = '□ 有'
        c = ws.cell(row=r, column=COL_AFTER + 3, value=interest)
        c.font = body_font
        c.alignment = center
        c.border = border

        # 備考
        c = ws.cell(row=r, column=COL_AFTER + 4, value=d['inventory_remarks'] if d else '')
        c.font = body_font
        c.alignment = left_wrap
        c.border = border

    # --- 合計行 ---
    total_r = DATA_START + MAX_ROWS
    # 「計」ラベル
    ws.merge_cells(start_row=total_r, start_column=1, end_row=total_r, end_column=COL_AFTER - 1)
    c = ws.cell(row=total_r, column=1, value='計')
    c.font = hdr_font
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = border
    for col in range(2, COL_AFTER):
        ws.cell(row=total_r, column=col).border = border

    total_pb = sum(r['passbook_balance'] or 0 for r in rows)
    c = ws.cell(row=total_r, column=COL_AFTER, value=total_pb)
    c.font = hdr_font
    c.alignment = right_a
    c.border = border
    c.number_format = '#,##0'

    ws.cell(row=total_r, column=COL_AFTER + 1).border = border

    total_cert = sum(r['certificate_balance'] or 0 for r in rows)
    c = ws.cell(row=total_r, column=COL_AFTER + 2, value=total_cert)
    c.font = hdr_font
    c.alignment = right_a
    c.border = border
    c.number_format = '#,##0'

    for col in range(COL_AFTER + 3, total_cols + 1):
        ws.cell(row=total_r, column=col).border = border

    # --- 列幅 ---
    widths = {'A': 4, 'B': 14, 'C': 12, 'D': 8, 'E': 12}
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w
    for i in range(year_count):
        ws.column_dimensions[get_column_letter(COL_YEAR_START + i)].width = 7
    after_widths = [10, 7, 10, 8, 16]
    for i, w in enumerate(after_widths):
        ws.column_dimensions[get_column_letter(COL_AFTER + i)].width = w

    # --- 印刷設定 ---
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f"{sanitize_filename(case.name)}_通帳有無一覧表.xlsx"
    set_download_filename(response, filename)
    return response


def api_reorder_passbook_inventory(request: HttpRequest, pk: int) -> JsonResponse:
    """通帳有無一覧表の並び順を保存（AJAX）"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    case = get_object_or_404(Case, pk=pk)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    order = data.get('order', [])
    if not order:
        return JsonResponse({'error': 'order required'}, status=400)

    for i, account_id in enumerate(order):
        Account.objects.filter(pk=account_id, case=case).update(print_order=i)

    return JsonResponse({'ok': True})
