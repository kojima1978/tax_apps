"""エクスポートビュー"""
import json
import logging
from datetime import datetime
from io import BytesIO

from django.contrib import messages
from django.db.models import Sum
from django.http import HttpRequest, HttpResponse
from django.shortcuts import get_object_or_404, redirect
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

from ..models import Case
from ..handlers import FIELD_LABELS, parse_amount
from ..lib import config
from ..lib.constants import sort_categories
from ..lib.text_utils import df_filter_by_keyword
from ..services import AnalysisService
from ..templatetags.japanese_date import wareki
from ._helpers import (
    sanitize_filename, set_download_filename, build_filter_state,
    build_filtered_filename, require_transactions, prepare_export_df,
    build_csv_response, get_export_columns,
)

logger = logging.getLogger(__name__)

# エクスポートタイプ設定: (フィルタフィールド, ファイル名サフィックス)
_EXPORT_TYPE_CONFIG = {
    'transfers': ('is_transfer', '資金移動'),
    'flagged':   ('is_flagged',  '付箋付き取引'),
    'all':       (None,          '全取引'),
}


def export_json(request: HttpRequest, pk: int) -> HttpResponse:
    """案件データをJSONでバックアップエクスポート"""
    logger.info(f"JSONエクスポート開始: case_id={pk}")
    case = get_object_or_404(Case, pk=pk)
    transactions = case.transactions.all().order_by('date', 'id')

    empty_redirect = require_transactions(request, transactions, pk, 'analysis-dashboard')
    if empty_redirect:
        return empty_redirect

    totals = transactions.aggregate(total_in=Sum('amount_in'), total_out=Sum('amount_out'))

    # 口座情報をアノテーションして取得
    transactions_with_account = transactions.with_account_info()
    export_fields = [
        'date', 'bank_name', 'branch_name', 'account_type', 'account_number',
        'description', 'amount_out', 'amount_in', 'balance',
        'category', 'holder', 'is_large', 'is_transfer', 'transfer_to',
        'is_flagged', 'memo',
    ]
    transactions_data = []
    for tx_dict in transactions_with_account.values(*export_fields):
        if tx_dict['date']:
            tx_dict['date'] = tx_dict['date'].isoformat()
        transactions_data.append(tx_dict)

    export_data = {
        'version': '1.0',
        'exported_at': datetime.now().isoformat(),
        'case': {
            'name': case.name,
            'created_at': case.created_at.isoformat() if case.created_at else None,
        },
        'transactions': transactions_data,
        'statistics': {
            'total_transactions': len(transactions_data),
            'total_in': totals['total_in'] or 0,
            'total_out': totals['total_out'] or 0,
        },
        'settings': config.load_user_settings(),
    }

    response = HttpResponse(
        json.dumps(export_data, ensure_ascii=False, indent=2),
        content_type='application/json; charset=utf-8'
    )
    filename = f"{sanitize_filename(case.name)}_backup.json"
    set_download_filename(response, filename)

    logger.info(f"JSONエクスポート完了: case_id={pk}, transactions={len(transactions_data)}")
    return response


def export_csv(request: HttpRequest, pk: int, export_type: str) -> HttpResponse:
    """取引データをCSVでエクスポート"""
    logger.info(f"CSVエクスポート開始: case_id={pk}, type={export_type}")
    case = get_object_or_404(Case, pk=pk)
    transactions = case.transactions.all().order_by('date', 'id')

    empty_redirect = require_transactions(request, transactions, pk)
    if empty_redirect:
        return empty_redirect

    df = pd.DataFrame(list(transactions.with_account_info().values()))

    config_entry = _EXPORT_TYPE_CONFIG.get(export_type)
    if config_entry:
        filter_field, suffix = config_entry
        if filter_field:
            df = df[df[filter_field]].copy()
        filename = f"{sanitize_filename(case.name)}_{suffix}.csv"
    else:
        filename = f"{sanitize_filename(case.name)}_取引データ.csv"

    if df.empty:
        messages.warning(request, "該当するデータがありません。")
        return redirect('analysis-dashboard', pk=pk)

    return build_csv_response(df, filename, include_memo=(export_type == 'flagged'))


def export_csv_filtered(request: HttpRequest, pk: int) -> HttpResponse:
    """絞り込み条件付きでCSVエクスポート"""
    logger.info(f"絞り込みCSVエクスポート開始: case_id={pk}")
    case = get_object_or_404(Case, pk=pk)

    filter_state = build_filter_state(request)
    transactions = case.transactions.with_account_info().order_by('date', 'id')
    transactions = AnalysisService.apply_filters(transactions, filter_state)

    amount_min_val, amount_min_ok = parse_amount(filter_state['amount_min']) if filter_state['amount_min'] else (None, True)
    amount_max_val, amount_max_ok = parse_amount(filter_state['amount_max']) if filter_state['amount_max'] else (None, True)
    amount_min = amount_min_val if amount_min_ok and amount_min_val else None
    amount_max = amount_max_val if amount_max_ok and amount_max_val else None

    empty_redirect = require_transactions(request, transactions, pk)
    if empty_redirect:
        return empty_redirect

    df = pd.DataFrame(list(transactions.values()))

    keyword = filter_state.get('keyword', '')
    if keyword:
        df = df_filter_by_keyword(df, keyword)
        if df.empty:
            messages.warning(request, "エクスポートするデータがありません。")
            return redirect('analysis-dashboard', pk=pk)

    filename = build_filtered_filename(case.name, filter_state, amount_min, amount_max)

    logger.info(f"絞り込みCSVエクスポート完了: case_id={pk}, count={len(df)}")
    return build_csv_response(df, filename, include_memo=True)


def export_xlsx_by_category(request: HttpRequest, pk: int) -> HttpResponse:
    """分類別にシート分けしたExcelファイルをエクスポート"""
    case = get_object_or_404(Case, pk=pk)
    transactions = case.transactions.all().order_by('date', 'id')

    empty_redirect = require_transactions(request, transactions, pk)
    if empty_redirect:
        return empty_redirect

    df = pd.DataFrame(list(transactions.with_account_info().values()))

    if 'date' in df.columns:
        df['date'] = df['date'].apply(lambda d: wareki(d, 'short'))

    cols_to_export, headers = get_export_columns(df.columns, exclude_balance=True)
    flagged_cols, flagged_headers = get_export_columns(df.columns, include_memo=True, exclude_balance=True)

    grouped = df.groupby('category')
    sorted_cats = sort_categories(grouped.groups.keys())

    # 金額カラムのインデックスを特定（1-based, ヘッダー行の次から）
    amount_fields = {'amount_out', 'amount_in'}

    def _format_sheet(ws, columns):
        """金額カラムにカンマ区切り書式を適用 + A4縦・列幅1ページの印刷設定"""
        # カンマ区切り書式
        amount_col_indices = [i + 1 for i, c in enumerate(columns) if c in amount_fields]
        for col_idx in amount_col_indices:
            col_letter = get_column_letter(col_idx)
            for cell in ws[col_letter]:
                if cell.row > 1:
                    cell.number_format = '#,##0'
        # 印刷設定: A4縦、列幅を1ページに収める
        ws.page_setup.paperSize = 9  # A4
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    wb = Workbook()
    wb.remove(wb.active)

    for cat in sorted_cats:
        cat_df = grouped.get_group(cat)
        ws = wb.create_sheet(title=str(cat)[:31])
        ws.append(headers)
        for _, row in cat_df[cols_to_export].iterrows():
            ws.append([row[c] for c in cols_to_export])
        _format_sheet(ws, cols_to_export)

    # 多額取引シートを追加
    threshold = config.load_user_settings().get('LARGE_AMOUNT_THRESHOLD', 500000)
    large_df = df[(df['amount_out'] >= threshold) | (df['amount_in'] >= threshold)]
    if not large_df.empty:
        threshold_label = f"{threshold // 10000}万円以上"
        ws = wb.create_sheet(title=threshold_label[:31])
        ws.sheet_properties.tabColor = 'DC3545'
        ws.append(headers)
        for _, row in large_df[cols_to_export].iterrows():
            ws.append([row[c] for c in cols_to_export])
        _format_sheet(ws, cols_to_export)

    # 付箋付き取引シートを末尾に追加
    flagged_df = df[df['is_flagged'] == True]  # noqa: E712
    if not flagged_df.empty:
        ws = wb.create_sheet(title='付箋付き')
        ws.sheet_properties.tabColor = 'FF8C00'
        ws.append(flagged_headers)
        for _, row in flagged_df[flagged_cols].iterrows():
            ws.append([row[c] for c in flagged_cols])
        _format_sheet(ws, flagged_cols)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f"{sanitize_filename(case.name)}_分類別取引.xlsx"
    set_download_filename(response, filename)
    return response
